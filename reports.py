import sqlite3
import pandas as pd
from datetime import datetime
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

def generate_styled_report():
    try:
        # 1. الاتصال وقراءة البيانات
        conn = sqlite3.connect('school_data.db')
        df = pd.read_sql_query("SELECT * FROM attendance", conn)
        
        if df.empty:
            print("⚠️ لا توجد بيانات لتصديرها!")
            return

        date_str = datetime.now().strftime("%Y-%m-%d")
        filename = f"تقرير_مدرسة_الأسد_الملون_{date_str}.xlsx"
        
        # 2. حفظ الملف بصيغة إكسل أولاً
        df.to_excel(filename, index=False)
        
        # 3. فتح الملف لإضافة التنسيقات (الألوان)
        wb = load_workbook(filename)
        ws = wb.active
        
        # تعريف الألوان (أخضر فاتح للحضور، أحمر فاتح للغياب)
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        # تنسيق رأس الجدول (Header)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # تلوين الصفوف بناءً على الحالة
        for row in range(2, ws.max_row + 1):
            status = ws.cell(row=row, column=3).value # عمود الحالة (status)
            for cell in ws[row]:
                cell.alignment = Alignment(horizontal="center")
                if status == "حاضر":
                    cell.fill = green_fill
                elif status == "غائب":
                    cell.fill = red_fill

        # حفظ التعديلات النهائية
        wb.save(filename)
        
        # 4. نقل الملف لمجلد التحميلات
        os.system(f"cp {filename} ~/storage/downloads/")
        
        print("\n" + "="*40)
        print(f"✅ تم إنشاء التقرير الملون بنجاح!")
        print(f"📂 اذهب لمجلد Downloads ستجد: {filename}")
        print("="*40)
        
        conn.close()
    except Exception as e:
        print(f"❌ حدث خطأ أثناء التلوين: {e}")

if __name__ == "__main__":
    generate_styled_report()



